from yaml import load, dump
from os import stat, environ
from platform import system
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from time import time
from google.cloud import storage


def get_zones_from_yaml(file_location="./zones.yaml", time_threshold=None):
    if not time_threshold:
        time_threshold = round(time()) - (3600 * 24 * 7)

    path = Path(file_location)
    if path.is_file():
        info = stat(file_location)
        if round(info.st_mtime) < time_threshold:
            return None  # force a refresh if file is too old
        with open(file_location) as file:
            return load(file, Loader=yaml.FullLoader)

    return {}


def write_to_yaml(data, file_location=None):
    if not file_location:
        file_location = "./" + data.lower() + ".yaml"
    path = Path(file_location)
    if path.is_file():
        info = stat(file_location)

    with open(file_location, 'w') as file:
        output_file = dump(data, file)

    return None


def read_from_bucket(bucket_name:str, file_name: str) -> list:

    try:
        storage_client = storage.Client()
        file_size: int = storage_client.bucket(bucket_name).get_blob(file_name).size
        file_text: str = storage_client.bucket(bucket_name).blob(file_name).download_as_text()
        if file_size > 0:
            return file_text.rstrip().splitlines()
        return []

    except Exception as e:
        raise e


def write_to_excel(sheets: dict = {'Sheetz': []}, excelfile: str = "Book1.xlsx", start_row: int = 1, start_col: int = 1):

    wb = Workbook()
    for sheet_name, data in sheets.items():

        # Create worksheet
        ws = wb.create_sheet(sheet_name)

        if len(data) < 1:
            continue
        if not isinstance(data[0], dict):
            continue


        # Write field names in the first row
        num_columns = 0
        column_widths = {}
        #print(data[0])
        for column_index, column_name in enumerate(data[0].keys()):
            #print(column_index, column_name)
            ws.cell(row=start_row, column=column_index + 1).value = column_name
            num_columns += 1
            column_widths[column_index] = len(str(column_name))

        # Write out rows of data
        for row_num in range(len(data)):
            #print(data[row_num])
            row = list(data[row_num].values())
            ws.append(row)

            # Keep track of largest value for each column
            for column_index, entry in enumerate(row):
                #print(column_index, entry)
                column_width = len(str(entry)) if entry else 0
                if column_index in column_widths:
                    if column_width > column_widths[column_index]:
                        column_widths[column_index] = column_width

        for i in range(num_columns):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_widths[i] + 1

    # Save the file
    wb.save(filename=excelfile)


def get_home_dir() -> str:

    my_os = system().lower()
    if my_os.startswith("win"):
        home_dir = environ.get("USERPROFILE")
        seperator = "\\"
    elif my_os:
        home_dir = environ.get("HOME")
        seperator = "/"

    return home_dir + seperator
